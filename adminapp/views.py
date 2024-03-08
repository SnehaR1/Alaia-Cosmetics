from .models import ProductVariant
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from userauth.models import CustomUser
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.utils import timezone
from django.db.models import Count, F
from .models import (
    Category,
    Brand,
    Product,
    ProductVariant,
    ProductImages,
    Color,
    Quantity,
    Order,
    OrderItem,
    Coupon,
    Offers,
    Blogs,
    BlogAdditionalImage,
    Wallet,
    ReturnedProduct,
)
from adminapp.forms import (
    UserEditForm,
    AddProduct,
    ProductVariantForm,
    ProductImagesForm,
    OrderItemForm,
    CouponForm,
    OffersForm,
    BlogForm,
    BlogImagesForm,
)
from datetime import timedelta
from django.db.models import Sum
from django.db.models.functions import TruncWeek, TruncMonth, TruncYear
from django.http import JsonResponse
from django.db.models.functions import ExtractYear, ExtractMonth, ExtractWeek
import calendar
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from django.core.mail import send_mail


def admin_login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = authenticate(request, username=email, password=password)

        if user is not None:
            if user.is_staff:
                login(request, user)

                return redirect("admin_dashboard")
            else:
                messages.error(request, "This account does not have admin privileges.")
                return render(request, "admin/admin_login.html")
        else:
            messages.error(request, "Incorrect Username or Password")
            return render(request, "admin/admin_login.html")
    return render(request, "admin/admin_login.html")


@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    filter_type = "weekly"
    if request.method == "POST":
        filter_type = request.POST.get("filter")
    orders = Order.objects.all()
    order_items = OrderItem.objects.all()
    products = Product.objects.all()
    total_revenue = Order.objects.aggregate(Sum("total_price")).get(
        "total_price__sum", 0
    )
    total_products_sold = OrderItem.objects.aggregate(Sum("quantity")).get(
        "quantity__sum", 0
    )
    total_customers = CustomUser.objects.aggregate(Count("email")).get(
        "email__count", 0
    )
    product_orders = (
        OrderItem.objects.annotate(product_title=F("product__title"))
        .values("product_title")
        .annotate(count=Count("id"))
        .annotate(price=F("price__old_price"))
        .annotate(amount=(F("count") * F("price")))
        .annotate(status=F("product__status"))
        .values("product_title", "count", "price", "amount", "status")
    )
    products = Product.objects.all()
    product_labels = []
    count = []
    for product in product_orders:
        product_labels.append(product["product_title"])
        count.append(product["count"])
    monthly_orders = (
        Order.objects.annotate(month=ExtractMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .values("month", "count")
    )

    month_labels = []
    total_monthly_orders = []
    for o in monthly_orders:
        month_labels.append(calendar.month_name[o["month"]])
        total_monthly_orders.append(o["count"])

    weekly_orders = (
        Order.objects.annotate(week=ExtractWeek("created_at"))
        .values("week")
        .annotate(count=Count("id"))
        .values("week", "count")
    )

    week_labels = []
    total_weekly_orders = []
    for o in weekly_orders:
        week_labels.append(f"Week {o['week']}")
        total_weekly_orders.append(o["count"])

    yearly_orders = (
        Order.objects.annotate(year=ExtractYear("created_at"))
        .values("year")
        .annotate(count=Count("id"))
        .values("year", "count")
    )

    year_labels = []
    total_yearly_orders = []
    for o in yearly_orders:
        year_labels.append(str(o["year"]))
        total_yearly_orders.append(o["count"])
    total_price = Order.objects.aggregate(Sum("total_price"))
    print(total_price)
    return render(
        request,
        "admin/admin_dashboard.html",
        {
            "product_orders": product_orders,
            "products": products,
            "orders": orders,
            "order_items": order_items,
            "total_products_sold": total_products_sold,
            "total_revenue": total_revenue,
            "total_customers": total_customers,
            "month_orders": monthly_orders,
            "month_labels": month_labels,
            "total_monthly_orders": total_monthly_orders,
            "week_orders": weekly_orders,
            "week_labels": week_labels,
            "total_weekly_orders": total_weekly_orders,
            "year_orders": yearly_orders,
            "year_labels": year_labels,
            "total_yearly_orders": total_yearly_orders,
            "filter_type": filter_type,
            "total_price": total_price,
            "product_labels": product_labels,
            "count": count,
        },
    )


@user_passes_test(lambda u: u.is_superuser)
def admin_users(request):
    users = CustomUser.objects.all()

    if request.method == "POST":
        form = UserEditForm(request.POST)
        if form.is_valid():
            user_id_to_block = form.cleaned_data.get("user_id_to_block")
            user_to_block = get_object_or_404(CustomUser, pk=user_id_to_block)
            user_to_block.is_active = not user_to_block.is_active
            user_to_block.save()
    else:
        form = UserEditForm()

    return render(request, "admin/admin_users.html", {"users": users, "form": form})


@user_passes_test(lambda u: u.is_superuser)
def admin_category(request):
    if request.method == "POST":
        category_name = request.POST.get("category_name")
        if category_name:
            Category.objects.create(title=category_name)
        return redirect("admin_category")

    categories = Category.objects.annotate(product_count=Count("product_set"))
    return render(request, "admin/admin_category.html", {"categories": categories})


@user_passes_test(lambda u: u.is_superuser)
def admin_brand(request):
    if request.method == "POST":
        brand_name = request.POST.get("brand_name")
        if brand_name:
            Brand.objects.create(title=brand_name)
        return redirect("admin_brand")

    brands = Brand.objects.annotate(product_count=Count("product_set"))
    return render(request, "admin/admin_brand.html", {"brands": brands})


@user_passes_test(lambda u: u.is_superuser)
def admin_products(request):
    products = Product.objects.all().order_by("-date")
    return render(request, "admin/admin_products.html", {"products": products})


@user_passes_test(lambda u: u.is_superuser)
def add_product(request):
    if request.method == "POST":
        form = AddProduct(request.POST)

        if form.is_valid():
            form.save()

            messages.success(request, f"Product added successfully!")
            return redirect("admin_products")
    else:
        form = AddProduct()

    context = {"form": form}

    return render(request, "admin/add_product.html", context)


@user_passes_test(lambda u: u.is_superuser)
def edit_product(request, pid=None):
    product = get_object_or_404(Product, pk=pid)
    if request.method == "POST":
        form = AddProduct(request.POST, request.FILES, instance=product)

        if form.is_valid():
            form.save()

            messages.success(request, f"Product Updated successfully!")
            return redirect("admin_products")
    else:
        form = AddProduct(
            instance=product,
            initial={
                "brand": product.brand,
                "category": product.category,
            },
        )

    context = {"form": form, "product": product}

    return render(request, "admin/edit_product.html", context)


@user_passes_test(lambda u: u.is_superuser)
def admin_variant(request):
    products = ProductVariant.objects.all().order_by("-date")

    context = {
        "products": products,
    }
    return render(request, "admin/admin_variants.html", context)


@user_passes_test(lambda u: u.is_superuser)
def add_variant(request):
    if request.method == "POST":
        form = ProductVariantForm(request.POST, request.FILES)
        img_form = ProductImagesForm(request.POST, request.FILES)

        if form.is_valid() and img_form.is_valid():
            product_variant = form.save(commit=False)

            new_color_name = request.POST.get("new_color")
            if new_color_name:
                color, created = Color.objects.get_or_create(name=new_color_name)
                product_variant.color = color
            else:
                product_variant.color = None

            new_quantity_name = request.POST.get("new_quantity")
            unit = request.POST.get("unit")

            if new_quantity_name:
                quantity, created = Quantity.objects.get_or_create(
                    name=new_quantity_name, unit=unit
                )
                product_variant.quantity = quantity
            else:
                product_variant.quantity = None  # or set it to another default value

            product_variant.save()

            images = request.FILES.getlist("images")

            for image in images:
                ProductImages.objects.create(
                    images=image, productvariant=product_variant
                )

            messages.success(request, "Product and images added successfully!")
            return redirect("admin_variant")
        else:
            error_message = "Failed to add the Offer! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)
    else:
        form = ProductVariantForm()
        img_form = ProductImagesForm()

    context = {
        "form": form,
        "img_form": img_form,
    }
    return render(request, "admin/add_variant.html", context)


def logout_view(request):
    logout(request)
    messages.success(request, "Logged out successfully!")
    return redirect("admin_login")


def edit_user(request, user_id):
    user = CustomUser.objects.get(pk=user_id)
    if request.method == "POST":
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            user.save()
            return redirect("admin_users")
    else:
        return redirect("admin_users")


@user_passes_test(lambda u: u.is_superuser)
def admin_orders(request):
    orders = Order.objects.all().order_by("-created_at")
    orderitems = OrderItem.objects.all()
    productvariants = Product.objects.select_related("productvariant").all()
    return render(
        request,
        "admin/admin_orders.html",
        {
            "orders": orders,
            "productvairants": productvariants,
            "orderitems": orderitems,
        },
    )


def update_status(request, order_item_pk):
    orders = Order.objects.all()
    orderitems = OrderItem.objects.all()
    productvariants = Product.objects.select_related("productvariant").all()
    order_item = OrderItem.objects.get(pk=order_item_pk)
    return_req = ReturnedProduct.objects.filter(order_id=order_item_pk).first()
    print(order_item_pk)
    if request.method == "POST":
        new_status = request.POST.get("status")
        print(new_status)
        if order_item.status == new_status:
            messages.error(request, "Status remains unchanged; already up to date.")
            return redirect("admin_orders")
        elif new_status == "request processing" or new_status == "Request Approved":
            messages.error(request, "invalid update")
            return redirect("admin_orders")
        elif new_status == "delivered":
            order_item.status = new_status
            order_item.delivery_date = timezone.now()
            order_item.save()
            messages.success(request, "Status updated successfully!!!")
            return redirect("admin_orders")
        elif new_status == "returned":
            if return_req and return_req.is_approved:
                order_item.status = new_status
                order_item.save()
                user = order_item.order.user
                wallet = Wallet.objects.get(user=user)
                wallet.balance += order_item.order.total_price
                wallet.save()
                subject = "Refund For Product Return"
                message = f"Dear {order_item.order.user.first_name},\n\nYour order for  {order_item.product.title} has been returned.The refunded amount of Rs {order_item.order.total_price} has been added to your wallet."
                from_email = "sneharavindranathan@gmail.com"
                recipient_list = [order_item.order.user.email]
                send_mail(subject, message, from_email, recipient_list)
                messages.success(request, "User notified with wallet refund.")
                return redirect("admin_orders")
            elif return_req and not return_req.is_approved:
                messages.success(
                    request, "Status update requires user request approval "
                )
                return redirect("admin_orders")
            else:
                messages.error(request, "Status update requires explicit user request")
                return redirect("admin_orders")

        elif (
            new_status == "cancelled"
            and order_item.order.payment_method != "cash_on_delivery"
        ):
            order_item.status = new_status
            order_item.save()
            user = order_item.order.user

            wallet = Wallet.objects.get(user=user)
            wallet.balance += order_item.order.total_price
            wallet.save()
            subject = "Your Order has been Cancelled"
            message = f"Dear {order_item.order.user.first_name},\n\nYour order for  {order_item.product.title} has been cancelled by the admin.The refunded amount of Rs {order_item.order.total_price} has been added to your wallet."
            from_email = "sneharavindranathan@gmail.com"
            recipient_list = [order_item.order.user.email]

            send_mail(subject, message, from_email, recipient_list)
            messages.success(
                request,
                "Order cancellation email sent.User notified with wallet refund.",
            )
            return redirect("admin_orders")
        elif (
            new_status == "cancelled"
            and order_item.order.payment_method == "cash_on_delivery"
        ):
            order_item.status = new_status
            order_item.save()
            subject = "Your Order has been Cancelled"
            message = f"Dear {order_item.order.user.first_name},\n\nYour order for  {order_item.product.title} has been cancelled by the admin."
            from_email = "sneharavindranathan@gmail.com"
            recipient_list = [order_item.order.user.email]

            send_mail(subject, message, from_email, recipient_list)
            messages.success(request, "Order cancellation email sent")
            return redirect("admin_orders")
        else:
            order_item.status = new_status
            order_item.save()
            messages.success(request, "Status updated successfully!!!")
            return redirect("admin_orders")

    return render(
        request,
        "admin/admin_orders.html",
        {
            "orders": orders,
            "productvairants": productvariants,
            "orderitems": orderitems,
        },
    )


def toggle_listing(request, pv_id):
    product_variant = get_object_or_404(ProductVariant, id=pv_id)

    if request.method == "POST":
        is_listed = request.POST.get("listing_status") == "listed"
        print(is_listed)
        product_variant.is_listed = is_listed
        product_variant.save()

    return render(
        request, "admin/admin_variants.html", {"product_variant": product_variant}
    )


def unlist_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == "POST":
        is_listed = request.POST.get("listing_status") == "listed"
        category.is_listed = is_listed
        category.save()
        return render(request, "admin/admin_category.html", {"category": category})


def unlist_brand(request, brand_id):
    brand = get_object_or_404(Brand, id=brand_id)
    if request.method == "POST":
        is_listed = request.POST.get("listing_status") == "listed"
        brand.is_listed = is_listed
        brand.save()
        return render(request, "admin/admin_brand.html", {"brand": brand})


def product_listing(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == "POST":
        is_listed = request.POST.get("listing_status") == "listed"
        product.is_listed = is_listed
        product.save()
        return render(request, "admin/admin_products.html", {"product": product})
    return render(request, "admin/admin_products.html", {"product": product})


@user_passes_test(lambda u: u.is_superuser)
def edit_variant(request, pv_id):
    product_variant = get_object_or_404(ProductVariant, id=pv_id)

    if request.method == "POST":
        form = ProductVariantForm(request.POST, request.FILES, instance=product_variant)
        img_form = ProductImagesForm(
            request.POST, request.FILES, instance=product_variant
        )

        if form.is_valid() and img_form.is_valid():
            product_variant = form.save(commit=False)

            new_color_name = request.POST.get("new_color")
            if new_color_name:
                color, created = Color.objects.get_or_create(name=new_color_name)
                product_variant.color = color

            new_quantity_name = request.POST.get("new_quantity")
            unit = request.POST.get("unit")
            if new_quantity_name:
                quantity, created = Quantity.objects.get_or_create(
                    name=new_quantity_name, unit=unit
                )
                product_variant.quantity = quantity

            product_variant.save()

            images = request.FILES.getlist("images")
            if images:
                ProductImages.objects.filter(productvariant=product_variant).delete()

            for image in images:
                ProductImages.objects.create(
                    images=image, productvariant=product_variant
                )

            messages.success(request, "Product variant updated successfully!")
            return redirect("admin_variant")

        for field, errors in form.errors.items():
            messages.error(request, f"Form error in {field}: {errors.as_text()}")

        for field, errors in img_form.errors.items():
            messages.error(request, f"Form error in {field}: {errors.as_text()}")

    else:
        form = ProductVariantForm(instance=product_variant)
        form.initial["product"] = product_variant.product
        img_form = ProductImagesForm(instance=product_variant)

    context = {
        "form": form,
        "img_form": img_form,
        "product_variant": product_variant,
    }

    return render(request, "admin/edit_variant.html", context)


@user_passes_test(lambda u: u.is_superuser)
def admin_coupon(request):
    coupons = Coupon.objects.all().order_by("valid_to")
    return render(request, "admin/admin_coupon.html", {"coupons": coupons})


@user_passes_test(lambda u: u.is_superuser)
def add_coupon(request):
    if request.method == "POST":
        form = CouponForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Coupon added Successfully!!!")
            return redirect("admin_coupon")
        else:
            error_message = "Failed to add the Coupon! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)
    else:
        form = CouponForm()

    return render(request, "admin/add_coupon.html", {"form": form})


@user_passes_test(lambda u: u.is_superuser)
def edit_coupon(request, c_id):
    coupon = get_object_or_404(Coupon, id=c_id)
    if request.method == "POST":
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            updated_coupon = form.save(commit=False)
            updated_coupon.save()
            messages.success(request, "Coupon edited Successfully!!!")
            return redirect("admin_coupon")
        else:
            error_message = "Failed to edit the Coupon! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)
    else:
        form = CouponForm(instance=coupon)

    return render(request, "admin/edit_coupon.html", {"form": form})


@user_passes_test(lambda u: u.is_superuser)
def admin_offers(request):
    offers = Offers.objects.all().order_by("valid_to")
    return render(request, "admin/admin_offers.html", {"offers": offers})


@user_passes_test(lambda u: u.is_superuser)
def add_offers(request):
    if request.method == "POST":
        form = OffersForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Offer Added Successfully")
            return redirect("admin_offers")
        else:
            error_message = "Failed to add the Offer! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)
    else:
        form = OffersForm()
    return render(request, "admin/add_offer.html", {"form": form})


def report(request):
    orders = Order.objects.all()
    order_items = OrderItem.objects.all()
    products = Product.objects.all()
    total_revenue = Order.objects.aggregate(Sum("total_price")).get(
        "total_price__sum", 0
    )
    total_products_sold = OrderItem.objects.aggregate(Sum("quantity")).get(
        "quantity__sum", 0
    )
    total_customers = CustomUser.objects.aggregate(Count("email")).get(
        "email__count", 0
    )
    product_orders = (
        OrderItem.objects.annotate(product_title=F("product__title"))
        .values("product_title")
        .annotate(count=Count("id"))
        .annotate(price=F("price__old_price"))
        .annotate(amount=(F("count") * F("price")))
        .annotate(status=F("product__status"))
        .values("product_title", "count", "price", "amount", "status")
    )
    products = Product.objects.all()
    product_labels = []
    count = []
    for product in product_orders:
        product_labels.append(product["product_title"])
        count.append(product["count"])
    monthly_orders = (
        Order.objects.annotate(month=ExtractMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .values("month", "count")
    )

    month_labels = []
    total_monthly_orders = []
    for o in monthly_orders:
        month_labels.append(calendar.month_name[o["month"]])
        total_monthly_orders.append(o["count"])

    weekly_orders = (
        Order.objects.annotate(week=ExtractWeek("created_at"))
        .values("week")
        .annotate(count=Count("id"))
        .values("week", "count")
    )

    week_labels = []
    total_weekly_orders = []
    for o in weekly_orders:
        week_labels.append(f"Week {o['week']}")
        total_weekly_orders.append(o["count"])

    yearly_orders = (
        Order.objects.annotate(year=ExtractYear("created_at"))
        .values("year")
        .annotate(count=Count("id"))
        .values("year", "count")
    )

    year_labels = []
    total_yearly_orders = []
    for o in yearly_orders:
        year_labels.append(str(o["year"]))
        total_yearly_orders.append(o["count"])
    total_price = Order.objects.aggregate(Sum("total_price"))
    template = get_template("admin/admin_dashboard_report.html")
    if request.GET.get("excel"):
        product_orders_for_excel = [
            {
                "product_title": item.get("product_title", ""),
                "count": item.get("count", ""),
                "price": item.get("price", ""),
                "amount": item.get("amount", ""),
                "status": "In stock" if item.get("status", False) else "Out of stock",
            }
            for item in product_orders
        ]
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        headers = ["Product name", "Sales", "Unit Price", "Amount", "Status"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        for row_num, item in enumerate(product_orders_for_excel, 2):
            sheet.cell(row=row_num, column=1, value=item["product_title"])
            sheet.cell(row=row_num, column=2, value=item["count"])
            sheet.cell(row=row_num, column=3, value=item["price"])
            sheet.cell(row=row_num, column=4, value=item["amount"])
            sheet.cell(row=row_num, column=5, value=item["status"])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=product_orders.xlsx"

        workbook.save(response)

        return response

    elif request.GET.get("pdf"):
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="sales_report.pdf"'

        context = {
            "product_orders": product_orders,
            "products": products,
            "orders": orders,
            "order_items": order_items,
            "total_products_sold": total_products_sold,
            "total_revenue": total_revenue,
            "total_customers": total_customers,
            "month_orders": monthly_orders,
            "month_labels": month_labels,
            "total_monthly_orders": total_monthly_orders,
            "week_orders": weekly_orders,
            "week_labels": week_labels,
            "total_weekly_orders": total_weekly_orders,
            "year_orders": yearly_orders,
            "year_labels": year_labels,
            "total_yearly_orders": total_yearly_orders,
            "total_price": total_price,
            "product_labels": product_labels,
            "count": count,
        }
        template_content = template.render(context)

        pdf_data = pisa.CreatePDF(template_content, dest=response)

        if pdf_data.err:
            return HttpResponse(
                "We had some errors <pre>" + template_content + "</pre>"
            )

        return response


@user_passes_test(lambda u: u.is_superuser)
def edit_offers(request, offer_id):
    offer = get_object_or_404(Offers, id=offer_id)
    if request.method == "POST":
        form = OffersForm(request.POST, instance=offer)
        if form.is_valid():
            updated_offer = form.save(commit=False)
            if not updated_offer.active:
                if updated_offer.product:
                    product_variants = updated_offer.product.product_variants.all()
                    for variant in product_variants:
                        variant.offer_price = None
                        variant.save()
                elif updated_offer.category:
                    products = updated_offer.category.product_set.all()
                    product_variants = ProductVariant.objects.filter(
                        product__in=products
                    )
                    for variant in product_variants:
                        variant.offer_price = None
                        variant.save()
            else:
                updated_offer.new_price()

            updated_offer.save()
            messages.success(request, "Offer edited Successfully!!!")
            return redirect("admin_offers")
        else:
            error_message = "Failed to edit the Coupon! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)
    else:
        form = OffersForm(
            instance=offer,
            initial={"product": offer.product, "category": offer.category},
        )

    return render(request, "admin/edit_offers.html", {"form": form})


@user_passes_test(lambda u: u.is_superuser)
def admin_blog(request):
    blogs = Blogs.objects.all().order_by("-date")
    return render(request, "admin/admin_blog.html", {"blogs": blogs})


@user_passes_test(lambda u: u.is_superuser)
def add_blog(request):
    if request.method == "POST":
        form = BlogForm(request.POST, request.FILES)
        img_form = BlogImagesForm(request.POST, request.FILES)
        if form.is_valid() and img_form.is_valid():
            blog_instance = form.save()
            # additional_image_instance = img_form.save(commit=False)
            # additional_image_instance.blog = blog_instance
            # additional_image_instance.save()
            images = request.FILES.getlist("add_images")
            for image in images:
                BlogAdditionalImage.objects.create(add_images=image, blog=blog_instance)

            messages.success(request, "Blog Added Successfully!!!")
            return redirect("admin_blog")
        else:
            print(form.errors)
            error_message = "Failed to add the Offer! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)
    else:
        form = BlogForm()
        img_form = BlogImagesForm()

    return render(request, "admin/add_blog.html", {"form": form, "img_form": img_form})


@user_passes_test(lambda u: u.is_superuser)
def edit_blog(request, blog_id):
    blog = get_object_or_404(Blogs, id=blog_id)

    if request.method == "POST":
        form = BlogForm(request.POST, request.FILES, instance=blog)
        img_form = BlogImagesForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            try:
                blog = form.save()

                add_images = request.FILES.getlist("add_images")

                if add_images:
                    BlogAdditionalImage.objects.filter(blog=blog).delete()

                for image in add_images:
                    BlogAdditionalImage.objects.create(add_images=image, blog=blog)

                messages.success(request, "Blog updated successfully!")
                return redirect("admin_blog")
            except Exception as e:
                print(f"Error: {e}")
                messages.error(request, "Something went wrong")
        else:
            print(form.errors)
            error_message = "Failed to add the Offer! Please correct the following errors: {}".format(
                form.errors
            )
            messages.error(request, error_message)

    else:
        form = BlogForm(instance=blog)
        img_form = BlogImagesForm(instance=blog)

    context = {
        "form": form,
        "img_form": img_form,
        "blog": blog,
    }
    return render(request, "admin/edit_blog.html", context)


@user_passes_test(lambda u: u.is_superuser)
def admin_request(request):
    return_reqs = ReturnedProduct.objects.all()
    if request.method == "POST":
        return_req_id = request.POST.get("reject_return_id")
        return_req = ReturnedProduct.objects.get(id=return_req_id)
        order_item = return_req.order
        if "Reject" in request.POST:
            subject = "Your Return Request has been Rejected"
            message = f"Dear {return_req.user.first_name},\n\nYour return request for  {order_item.product.title} has been rejected by the admin."
            from_email = "sneharavindranathan@gmail.com"
            recipient_list = [return_req.user.email]
            send_mail(subject, message, from_email, recipient_list)
            order_item.status = "delivered"
            order_item.save()
            return_req.delete()
            messages.success(request, "Return request rejected and email sent")
            return redirect("admin_request")
        elif "Approve" in request.POST:
            return_req.is_approved = True
            return_req.save()
            subject = "Your Return Request has been Approved"
            message = f"Dear {return_req.user.first_name},\n\nYour return request for  {order_item.product.title} has been Approved by the admin."
            from_email = "sneharavindranathan@gmail.com"
            recipient_list = [return_req.user.email]
            send_mail(subject, message, from_email, recipient_list)
            order_item.status = "Request Approved"
            order_item.save()
            messages.success(request, "Return request Approved and email sent")
            return redirect("admin_request")

    return render(request, "admin/admin_request.html", {"return_reqs": return_reqs})


def custom_404_view(request, exception):
    return render(request, "404.html", status=404)
